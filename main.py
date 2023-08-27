import streamlit as st
import time
from datetime import datetime,timedelta
import json
import urllib.request
import requests
import openpyxl
import ftplib
import string
import random
#hello
N = 5
HOSTNAME = "ftp.ziply.at"
USERNAME = "u891522487.ziply.at"
PASSWORD = "upa86g@Z"
path = ""
polltime=2
pollgap=2
u2=""
replyto="5560841599"
looper=0
message = ""
updatetext1=0
token="6213442263:AAE_YtjdAwqXCh4sk4upKkwQ0R6nBnhxsAM"
headers = {"accept": "application/json","content-type": "application/json"}
photourl="https://api.telegram.org/bot"+token+"/sendPhoto"
docurl="https://api.telegram.org/bot"+token+"/sendDocument"
texturl="https://api.telegram.org/bot"+token+"/sendMessage"
meurl ="https://api.telegram.org/bot"+token+"/getMe"
pollurl="https://api.telegram.org/bot"+token+"/getUpdates"
filedownload="https://api.telegram.org/file/bot"+token+"/"
fileurl="https://api.telegram.org/bot"+token+"/getfile"
healthmessage="This is health message of URL Shortener BOT. It is running fine and you are receiving this message every 30 minutes "
healthtime = int(time.time())

user = path + "ziply.xlsx"
wb_obj = openpyxl.load_workbook(user)
sheet_obj = wb_obj.active
while looper==0:
    cn = 0
    mime=""
    messagelink = ""
    updatefile = open(path + "updateid.txt", 'r+')
    updatetext = updatefile.read()
    updatefile.close()
    #Starting Long Polling
    lastupdate=int(updatetext)
    payloadpoll = {"offset":lastupdate,"limit": 50,"timeout": pollgap}
    response = requests.post(pollurl, json=payloadpoll, headers=headers)
    f = urllib.request.urlopen(pollurl)
    data=json.load(f)
    b=data

    c=str(b)

    lenc=len(c)
    print(c)
    print(lenc)
    if lenc<50:
        print("No New message "+str(datetime.now().strftime("%H:%M:%S")))
    if lenc>50:
      print("Message received at: "+str(datetime.now().strftime("%H:%M:%S")))
      tempstr = c[0:lenc]
      latid=""
      longid=""
      while cn==0:
       downloadlink=""
       usernametext=""
       usernamepos1=tempstr.find("username")
       if usernamepos1!=-1:
         usernamepos2=tempstr.find("language_code")
         usernametext=tempstr[usernamepos1+12:usernamepos2-4]
       filepos1 = tempstr.rfind("file_id':")
       if filepos1 != -1:

           filepos2 = tempstr.rfind("'file_unique_id':")
           fileid = tempstr[filepos1 + 11:filepos2 - 3]
           print(fileid)
           payload = {"file_id": fileid}
           response = requests.post(fileurl, json=payload, headers=headers)
           ok=response.text
           print(response.text)

           if  ok.find("true")!=-1:
               linkpos1 = ok.find("file_path")
               linkpos2 = ok.find("}}")
               link=ok[linkpos1 + 12:linkpos2 - 1]
               print(link)
               finddot= link.rfind(".")
               if finddot!=-1:
                   mime=link[finddot:len(link)]
                   print("mime:"+mime)
                   if mime != ".jpg" and mime != ".png" and mime != ".jpeg" and mime != ".bmp" and mime != ".webp" and mime =="":
                       messagelink="This is not a valid Image file. Please check and send Image file only"
               downloadlink=filedownload+link
               print(downloadlink)

       messagepostemp=0
       updatepos1 = tempstr.find("update_id':")
       print("updatepos1",updatepos1)
       if updatepos1==-1:
        break
       namepos = tempstr.find("is_bot': False, 'first_name':")
       namepos1 = tempstr.find("last_name")
       namepos2 = tempstr.find("username")
       namepos3 = tempstr.find("language_code")
       print("namepos", namepos)
       print("namepos1", namepos1)
       numberpos = tempstr.find("from")
       print("numberpos",numberpos)
       datepos = tempstr.find("date':")
       print("datepos",datepos)
       endpos1 = tempstr.find("}}, {")
       print("endpos1",endpos1)
       endpos2 = tempstr.find("}}]}")
       messagepos = tempstr.find("text")
       if messagepos-datepos>20:
         messagepostemp =messagepos
         messagepos=datepos+20
       print("messagepos",messagepos)
       updatetext1 = tempstr[updatepos1 + 12:updatepos1 + 21]
       print(updatetext1)
       numbertext = tempstr[numberpos + 14:numberpos + 24]
       print(numbertext)
       print(namepos1)

       if namepos1 == -1 :
           nametext = tempstr[namepos + 31:namepos3 - 4]

       if namepos1 != -1:
           nametext = tempstr[namepos + 31:namepos1 - 4]

       if namepos1 == -1 and namepos2!=-1 :
           nametext = tempstr[namepos + 31:namepos2 - 4]

       if namepos1 == -1 and namepos2 == -1 and namepos3!=-1:
           nametext = tempstr[namepos + 31:namepos3 - 4]

       print("Name: " + nametext)
       datetext = tempstr[datepos + 7:datepos + 17]
       dateconfirm = datetext[2:10]
       datetext=(datetime.fromtimestamp(int(datetext)).strftime('%Y-%m-%d %H:%M:%S'))
       print(datetext)
       if endpos1!=-1:
        messagetext=tempstr[messagepos+8:endpos1-1]
        if messagepostemp-datepos>20:
          messagetext="No Text"
       if endpos1==-1:
        messagetext=tempstr[messagepos+8:endpos2-1]
        if messagepostemp-datepos>20:
          messagetext="No Text"
       #print(messagetext)
       messagetext = messagetext.replace("\\xa0", " ")
       spcfind = messagetext.find("entities")
       if spcfind != -1:
           messagetext = messagetext[0:spcfind - 4]
       if messagetext[0:7] != "http://" and messagetext[0:8] != "https://" and messagetext[0:6]!="/start"  and messagetext!="sendfile" :
           message = "This is not a valid url or a weblink. Urls must start with http or https Please check and paste a valid url only "
           payloadtext = {
               "text": message,
               "parse_mode": "html",
               "disable_web_page_preview": True,
               "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
           response = requests.post(texturl, json=payloadtext, headers=headers)
           break
       if messagetext == "sendfile":
           file = "ziply.xlsx"
           files = {'document': open(file, 'rb')}
           response = requests.post(docurl + "?chat_id={}".format(numbertext), files=files)
           print(response.text)
           break
       print(messagetext)
       print("----------------")

       print("username: "+usernametext)

       apos=nametext.find(",")
       if apos!=-1:
           u1=nametext[0:apos-1]
       else:
        u1 = nametext

       if u1 != u2:
           alert = u1 + " has just logged in\n\n"
           alert1 = "<a href='tg://user?id=" + numbertext + "'>Click to chat here</a>"
           if usernametext!="":
               alert1=alert1+"\n\nor "+"@"+usernametext
           payloadtext = {"text": alert , "parse_mode": "html", "disable_web_page_preview": False,
                          "disable_notification": False, "reply_to_message_id": None, "chat_id": replyto}
           response = requests.post(texturl, json=payloadtext, headers=headers)
           u2 = u1
       callbackpos=messagetext.find("'data':")
       callbacktext=messagetext[callbackpos+9:len(messagetext)]
       #print("callbacktext:",callbacktext)
       numbertext=numbertext.rstrip(",")
       if messagetext[0:6] == "/start":
           payloadtext = {"text": "Hello "+u1 +"\n\nWelcome to Free URL Shortener Bot. Just paste any URL and the bot will shorten it instantly", "parse_mode": "html",
                          "disable_web_page_preview": True,
                          "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
           response = requests.post(texturl, json=payloadtext, headers=headers)
           break
       else:
               ftp_server = ftplib.FTP(HOSTNAME, USERNAME, PASSWORD)
               url = messagetext
               content = "<meta http-equiv=" + '"refresh"' + ' content="0; url=' +"'"+url +"'"+'"'+" />"
               #content1 = '<!DOCTYPE html> <html> <frameset cols="100% ,0%"> <frame noresize src=' + '"' + "index1.htm" + '"' + '> <frame src="https://eazyai.io">  </frameset> </html>'
               #print(content)
               #print(content1)
               file1 = open(path + "index.txt", "w")
               file1.write(content)
               file1.close()
               #file2 = open(path + "index1.txt", "w")
               #file2.write(content1)
               #file2.close()
               sr = 0
               sroww = 2
               while sr == 0:
                   res = ''.join(random.choices(string.ascii_lowercase +
                                                string.digits, k=N))
                   #print("res " + res)
                   sr1 = 0
                   srow = 2
                   while sr1 == 0:
                       cell_obj1 = sheet_obj.cell(row=srow, column=1)
                       cell_obj2 = sheet_obj.cell(row=srow, column=2)
                       cell_obj3 = sheet_obj.cell(row=srow, column=3)
                       cell_obj4 = sheet_obj.cell(row=srow, column=4)
                       cell_obj5 = sheet_obj.cell(row=srow, column=5)
                       if cell_obj3.value == res:
                           break
                       if cell_obj1.value == None or cell_obj1.value == "":
                           ftp_server.encoding = "utf-8"
                           ftpResponse = ftp_server.mkd(res)
                           ftp_server.cwd(res)
                           filename = "index.txt"
                           # filename1 = r"C:\Users\tapan\OneDrive\Desktop\urlshort\index1.txt"
                           with open(filename, "rb") as file:
                               ftp_server.storbinary('STOR ' + "index.htm", file)

                           # with open(filename1, "rb") as file:
                           # ftp_server.storbinary('STOR ' + "index1.htm", file)
                           #ftp_server.dir()
                           ftp_server.quit()
                           print("https://ziply.at/" + res)
                           print(len("https://ziply.at/" + res))
                           payloadtext = {
                               "text": "Your Shortened URL is :\n\n" + "https://ziply.at/" + res,
                               "parse_mode": "html",
                               "disable_web_page_preview": True,
                               "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
                           response = requests.post(texturl, json=payloadtext, headers=headers)
                           cell_obj1.value = numbertext
                           cell_obj2.value = datetext
                           cell_obj3.value = res
                           cell_obj4.value = "https://ziply.at/"+res
                           cell_obj5.value = messagetext
                           wb_obj.save(user)
                           break

                       srow = srow + 1
                   break
               break

       tempstr = tempstr[endpos1+4 :lenc]


    if int(updatetext1)>=int(updatetext):
      updatefile = open(path + "updateid.txt", 'w')
      updatetext = int(updatetext1)+1
      updatefile.write(str(updatetext))
      updatefile.close()

    healthtime1 = int(time.time())
    # print(healthtime1)
    if healthtime1 - healthtime > 1800:
        payloadtext = {"text": healthmessage, "parse_mode": "html", "disable_web_page_preview": False,
                       "disable_notification": False, "reply_to_message_id": None, "chat_id": replyto}
        response = requests.post(texturl, json=payloadtext, headers=headers)
        healthtime = healthtime1

    print("Offset ID in Text File updated succesfully")
